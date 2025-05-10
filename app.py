from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import logging

# Initialize Flask app
app = Flask(__name__)
CORS(app)  # Enable CORS

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# File paths
EXCEL_DIR = 'data'
EXCEL_FILE = os.path.join(EXCEL_DIR, 'submissions.xlsx')
os.makedirs(EXCEL_DIR, exist_ok=True)

def init_excel():
    """Initialize Excel file with headers"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        headers = [
            "Timestamp", "First Name", "Last Name", 
            "Email", "Phone", "Services", "Message"
        ]
        ws.append(headers)
        wb.save(EXCEL_FILE)
        logger.info("Created new Excel file")

@app.route('/')
def home():
    return render_template('form.html')

@app.route('/submit', methods=['POST'])
def submit():
    logger.info("\n=== New Form Submission ===")
    
    # Debug request data
    logger.info(f"Headers: {request.headers}")
    logger.info(f"Raw Data: {request.data}")
    
    if not request.is_json:
        logger.error("Request is not JSON")
        return jsonify({
            "success": False,
            "error": "Content-Type must be application/json"
        }), 400

    try:
        data = request.get_json()
        logger.info(f"Parsed JSON: {data}")
        
        # Validate required fields
        required = ['firstName', 'lastName', 'email']
        missing = [f for f in required if not data.get(f)]
        if missing:
            logger.error(f"Missing fields: {missing}")
            return jsonify({
                "success": False,
                "error": f"Missing required fields: {', '.join(missing)}",
                "required_fields": required
            }), 400

        # Prepare data for Excel
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        services = ", ".join(data.get('services', []))
        
        # Save to Excel
        init_excel()
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([
            timestamp,
            data['firstName'],
            data['lastName'],
            data['email'],
            data.get('phone', ''),
            services,
            data.get('message', '')
        ])
        wb.save(EXCEL_FILE)
        
        logger.info("Data saved successfully")
        return jsonify({
            "success": True,
            "message": "Submission received",
            "received_data": data  # For debugging
        })

    except Exception as e:
        logger.error(f"Error: {str(e)}", exc_info=True)
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

if __name__ == '__main__':
    init_excel()
    app.run(debug=True, host='0.0.0.0', port=5000)